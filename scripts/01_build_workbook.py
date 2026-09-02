#!/usr/bin/env python3
"""Baut die Ergebnis-Arbeitsmappe fuer die Verdichtungspotenzial-Analyse Thunersee.

Da alle im Auftrag genannten Datenquellen (Kanton-Bern-Geoportal, api3.geo.admin.ch,
opendata.swiss, regiogis.ch, ImmoScout24/Homegate/Comparis/RealAdvisor) in dieser
Ausfuehrungsumgebung durch die Netzwerk-Egress-Policy blockiert sind (siehe
logs/network_access_check.log), enthaelt Tabelle 1 KEINE Parzellen-Zeilen: es waeren
sonst geschaetzte/erfundene Werte noetig, was der Auftrag explizit verbietet.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"

wb = openpyxl.Workbook()

# ---------------------------------------------------------------------------
# Sheet 1: Ergebnis-Tabelle (Struktur gemaess Auftrag, Kapitel 6)
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Parzellen"

headers = [
    "Gemeinde", "Ort", "Strasse", "Parzellennummer", "Fläche m²",
    "Aktuelle Nutzungszone", "Verdichtungspotenzial (Begründung)",
    "Wer baut es (Baugesuch/Bauabsicht bekannt)", "Preis",
    "Name Besitzer", "Vorname Besitzer", "Auflagen",
    "m²-Preis Nachbarschaft (Ø ≥3 Referenzen)", "Besonderheiten",
    "GIS-Link (Permalink Parzelle)", "Quelle & Stand je Kernwert",
    "Vertrauensstatus",
]

header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(name=FONT_NAME, bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="top")

widths = [14, 12, 16, 16, 10, 18, 30, 22, 16, 16, 16, 20, 20, 22, 24, 30, 16]
for col_idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Platzhalter-Zeile: erklaert, warum keine Daten vorhanden sind (keine erfundenen Werte)
note = (
    "KEINE ZEILEN: Datenerhebung nicht möglich – alle benötigten Geodaten- und "
    "Vergleichspreis-Quellen waren in dieser Ausführungsumgebung netzwerkseitig "
    "blockiert (siehe Blatt 'Methodik & Quellen', Abschnitt 'Netzwerk-Zugriffsprüfung'). "
    "Es wurden bewusst keine geschätzten Werte eingetragen."
)
ws.cell(row=2, column=1, value=note).font = Font(name=FONT_NAME, italic=True, color="C00000")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 45

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

# ---------------------------------------------------------------------------
# Sheet 2: Methodik & Quellen
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Methodik & Quellen")
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 100

def write_row(r, a, b=None, bold_a=False, size_a=None, wrap=True):
    ca = ws2.cell(row=r, column=1, value=a)
    ca.font = Font(name=FONT_NAME, bold=bold_a, size=size_a or 11)
    ca.alignment = Alignment(wrap_text=wrap, vertical="top")
    if b is not None:
        cb = ws2.cell(row=r, column=2, value=b)
        cb.font = Font(name=FONT_NAME, size=11)
        cb.alignment = Alignment(wrap_text=wrap, vertical="top")
    return r + 1

r = 1
r = write_row(r, "Verdichtungspotenzial Parzellen Thunersee – Methodik & Quellen", bold_a=True, size_a=14)
r = write_row(r, "Stand", "2026-09-02 (UTC)")
r += 1

r = write_row(r, "0. ERGEBNIS IN KÜRZE", bold_a=True, size_a=12)
r = write_row(
    r,
    "Status",
    "Die eigentliche Parzellen-Auswertung (Schritte 1–11 des Auftrags) konnte NICHT "
    "durchgeführt werden. Grund: sämtliche im Auftrag genannten Datenquellen "
    "(kantonales Geoportal, amtliche Vermessung, opendata.swiss, RegioGIS, "
    "Immobilienportale) sind in dieser Claude-Code-Ausführungsumgebung durch die "
    "Netzwerk-Egress-Policy blockiert (HTTP 403 auf CONNECT-Ebene am Proxy-Gateway, "
    "nicht durch die Zielserver selbst). Es wurden gemäss Auftrag KEINE geschätzten "
    "oder erfundenen Werte eingetragen; Tabelle 'Parzellen' bleibt daher leer.",
)
r += 1

r = write_row(r, "1. NETZWERK-ZUGRIFFSPRÜFUNG (Schritt 0 des Auftrags)", bold_a=True, size_a=12)
r = write_row(
    r,
    "Testmethode",
    "curl (HTTP HEAD/GET, --max-time 15s) über den in dieser Umgebung vorgeschriebenen "
    "HTTPS-Proxy sowie zusätzlich das WebFetch-Tool als unabhängiger Pfad. Beide Wege "
    "liefern dasselbe Ergebnis. Rohdaten: logs/network_access_check.log "
    "(inkl. Proxy-Statusabfrage mit Einzel-Fehlermeldungen je Domain).",
)

domain_tests = [
    ("geo.apps.be.ch", "Kanton Bern Geoportal – Basiskarte Grundeigentum", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("agi.dij.be.ch", "Amt für Geoinformation Kanton Bern (MOPUBE/AV)", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("api3.geo.admin.ch", "Bundes-Geoportal / REST-API (geo.admin.ch)", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("map.geo.admin.ch", "Bundes-Geoportal Kartenviewer", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("opendata.swiss", "Open-Data-Portal Schweiz", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("regiogis.ch", "RegioGIS (Gemeinde-Zonenpläne Region Thun)", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("immoscout24.ch", "Vergleichsobjekte / Inserate", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("homegate.ch", "Vergleichsobjekte / Inserate", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("comparis.ch", "Vergleichsobjekte / Inserate", "blockiert (403 CONNECT, Gateway-Policy)"),
    ("realadvisor.ch", "Vergleichsobjekte / Verkaufsmeldungen", "blockiert (403 CONNECT, Gateway-Policy)"),
]
r += 1
hdr = ws2.cell(row=r, column=1, value="Domain")
hdr.font = Font(name=FONT_NAME, bold=True)
hdr2 = ws2.cell(row=r, column=2, value="Zweck / Ergebnis")
hdr2.font = Font(name=FONT_NAME, bold=True)
r += 1
for domain, purpose, result in domain_tests:
    r = write_row(r, domain, f"{purpose} — {result}")
r += 1
r = write_row(
    r,
    "Einordnung",
    "Die Proxy-Statusabfrage (__agentproxy/status) zeigt für jede der 10 Domains "
    "identisch: 'gateway answered 403 to CONNECT (policy denial or upstream failure)'. "
    "Dies ist eine Netzwerk-Policy-Blockade dieser Ausführungsumgebung, keine "
    "Bot-Abwehr/kein Login-Zwang der Zielseiten und kein TLS/Zertifikatsproblem. "
    "Gemäss Betriebsanleitung der Umgebung sind solche Policy-Denials nicht durch "
    "Wiederholung oder alternative Tools umgehbar.",
)
r += 1

r = write_row(r, "2. AUSWIRKUNG AUF DEN AUFTRAG", bold_a=True, size_a=12)
r = write_row(
    r, "Schritt 1 (Gemeindegrenzen)", "Nicht durchführbar – Geoportal-Zugriff blockiert.")
r = write_row(
    r, "Schritt 2 (Amtliche Vermessung: Parzellen/Gebäude)", "Nicht durchführbar – agi.dij.be.ch / api3.geo.admin.ch blockiert.")
r = write_row(
    r, "Schritt 3 (Zonenpläne)", "Nicht durchführbar – RegioGIS/Gemeinde-Geoportale blockiert.")
r = write_row(
    r, "Schritte 4–9 (Filterung, Gebäudetyp, Ausnützungsreserve, Eigentümerauskunft, Auflagen)",
    "Nicht durchführbar, da bereits die Grunddaten aus Schritt 2/3 fehlen.")
r = write_row(
    r, "Schritt 10 (Vergleichspreise)", "Nicht durchführbar – ImmoScout24/Homegate/Comparis/RealAdvisor blockiert.")
r = write_row(
    r, "Schritt 11 (Ergebnistabelle)", "Tabelle 'Parzellen' wurde mit korrekter Spaltenstruktur angelegt, aber bewusst ohne Dateninhalte (0 Zeilen) belassen.")
r = write_row(
    r, "Schritt 12 (Pilotlauf Thun/Gwatt)", "Konnte nicht gestartet werden, da bereits Schritt 0 (Netzwerkzugriff) fehlschlägt.")
r += 1

r = write_row(r, "3. QUALITÄTSREGEL EINGEHALTEN", bold_a=True, size_a=12)
r = write_row(
    r,
    "Keine Schätzwerte",
    "Es wurden gemäß Auftrag ('Keine geschätzten/erfundenen Zahlen', 'Bei Blockade "
    "dokumentieren statt mit Schätzwerten weiterarbeiten') keinerlei Platzhalter- oder "
    "Schätzdaten für Fläche, Zone, Preis oder Eigentümer eingetragen. Die Ergebnistabelle "
    "bleibt leer, bis echte Daten beschafft werden können.",
)
r += 1

r = write_row(r, "4. EMPFOHLENE NÄCHSTE SCHRITTE", bold_a=True, size_a=12)
r = write_row(
    r, "a) Netzwerk-Policy anpassen",
    "Die 10 oben genannten Domains müssten in der Egress-Policy dieser Ausführungsumgebung "
    "freigegeben werden (siehe /root/.ccr/README.md: 'organization policy denials' – nur ein "
    "Administrator/Owner der Umgebung kann dies ändern, z.B. über die "
    "Environment-/Netzwerk-Einstellungen dieser Claude-Code-Umgebung).")
r = write_row(
    r, "b) Alternative Ausführungsumgebung",
    "Den Auftrag in einer Umgebung mit freiem/anders konfiguriertem Internetzugang "
    "(z.B. lokale Ausführung auf einem Gerät mit normalem Internetzugang) erneut starten, "
    "damit die realen Geodaten- und Vergleichspreis-Abfragen möglich sind.")
r = write_row(
    r, "c) Nach Freigabe",
    "Sobald Netzwerkzugriff besteht: dieses Repo erneut aufrufen und Schritte 0–12 gemäss "
    "ursprünglicher Auftragsbeschreibung ausführen (Pilotlauf Thun/Gwatt zuerst, dann "
    "Ausweitung auf Spiez/Einigen, Hilterfingen/Hünibach, Oberhofen am Thunersee, Sigriswil/Gunten).")
r += 1

r = write_row(r, "5. TECHNISCHE ARTEFAKTE IN DIESEM REPO", bold_a=True, size_a=12)
r = write_row(r, "scripts/00_check_network_access.sh", "Führt die Erreichbarkeitsprüfung (Schritt 0) aus und schreibt logs/network_access_check.log.")
r = write_row(r, "logs/network_access_check.log", "Rohprotokoll der Verbindungstests inkl. Proxy-Gateway-Fehlermeldungen je Domain (Zeitstempel, HTTP-Status).")
r = write_row(r, "output/verdichtungspotenzial_thun.xlsx", "Diese Datei: Ergebnisstruktur (0 Zeilen) + Methodik-Dokumentation.")

ws2.freeze_panes = "A2"

import os

out_path = os.path.join(os.path.dirname(__file__), "..", "output", "verdichtungspotenzial_thun.xlsx")
wb.save(out_path)
print("saved", out_path)
